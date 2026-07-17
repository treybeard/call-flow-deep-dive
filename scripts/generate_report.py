#!/usr/bin/env python3
"""Generate multi-tab Excel report analyzing call flow data from CSV exports.

Reads summary CSVs from Reports/Charts/, generates an XLSX with multiple tabs
showing KPIs, monthly breakdowns, duration stats, transfer patterns, outcomes,
and call flow paths. All time data displays in MM:SS format.

Usage:
    cd scripts
    python3 generate_report.py
    # or specify category:
    python3 generate_report.py Switchboard
"""

import os
import sys
import csv
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

sys.path.insert(0, '/Library/Python/3.9/lib/python/site-packages')

PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CHARTS_DIR = os.path.join(PROJECT_DIR, "Reports", "Charts")
OUTPUT_DIR = os.path.join(PROJECT_DIR, "Output")

MONTH_ORDER = ['January', 'February', 'March', 'April', 'May', 'June',
               'July', 'August', 'September', 'October', 'November', 'December']


def format_duration(seconds):
    """Convert seconds to MM:SS string."""
    if seconds is None or seconds == 0:
        return "00:00"
    try:
        seconds = float(seconds)
    except (ValueError, TypeError):
        return "00:00"
    m = int(seconds) // 60
    s = int(seconds) % 60
    return f"{m:02d}:{s:02d}"


def parse_month_sort_key(month_name):
    """Return sort key for month name (lower case comparison)."""
    for i, m in enumerate(MONTH_ORDER):
        if month_name.lower() == m.lower():
            return i
    return 99


def get_ordered_months(rows_list):
    """Extract and sort months from outcome data rows."""
    months = []
    for row in rows_list:
        m = row.get('Month', '').strip()
        if m and m != 'Total' and m not in months:
            months.append(m)
    return sorted(months, key=parse_month_sort_key)


def read_csv_file(filepath):
    """Read a CSV file and return list of dicts."""
    rows = []
    if not os.path.exists(filepath):
        return rows
    with open(filepath, 'r', encoding='utf-8', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows


def write_header(ws, row, col, value, bold=True, fill=None):
    """Write a header cell with styling."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=11)
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    return cell


def write_cell(ws, row, col, value, wrap=False, bold=False):
    """Write a regular cell with border."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(horizontal='center', wrap_text=wrap)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    return cell


def create_summary_tab(wb, outcomes, duration, transfers, flow_paths):
    """Create the Summary tab with KPI cards and monthly comparison."""
    ws = wb.active
    ws.title = "Summary"

    # KPI Cards - row 4-8
    headers = ['Metric', '', 'Value', '', '']
    fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')

    # KPI title
    ws.merge_cells('A4:E4')
    kpi_title = ws['A4']
    kpi_title.value = 'Key Performance Indicators'
    kpi_title.font = Font(bold=True, size=14, color='FFFFFF')
    kpi_title.fill = fill
    kpi_title.alignment = Alignment(horizontal='center')

    # KPI values
    months = get_ordered_months(outcomes)
    kpi_rows = []

    # Total Calls
    total_calls = [r.get('Total', '0') for r in outcomes if r.get('Month', '').strip() in months]
    total_all = sum(int(t) for t in total_calls)
    kpi_rows.append(('Total Calls', '', total_all))

    # Calculate abandon rate
    total_abandoned = 0
    for m in months:
        for r in outcomes:
            if r.get('Month', '').strip() == m and r.get('Abandoned'):
                total_abandoned += int(r.get('Abandoned', 0))
    abandon_rate = round(total_abandoned / total_all * 100, 1) if total_all else 0
    kpi_rows.append(('Abandonment Rate', '', f'{abandon_rate}%'))

    # Success rate
    total_successful = 0
    for m in months:
        for r in outcomes:
            if r.get('Month', '').strip() == m and r.get('Successful'):
                total_successful += int(r.get('Successful', 0))
    success_rate = round(total_successful / total_all * 100, 1) if total_all else 0
    kpi_rows.append(('Success Rate', '', f'{success_rate}%'))

    # Avg Duration
    avg_dur_values = [r.get('Avg Duration (s)', '0') for r in duration if r.get('Month', '').strip() in months]
    avg_durs = [float(d) for d in avg_dur_values if d != '0']
    overall_avg = sum(avg_durs) / len(avg_durs) if avg_durs else 0
    kpi_rows.append(('Avg Duration', '', format_duration(overall_avg)))

    for i, (metric, _, val) in enumerate(kpi_rows):
        row = i + 5
        cell_label = ws.cell(row=row, column=1, value=metric)
        cell_label.font = Font(bold=True, size=11)
        cell_label.alignment = Alignment(horizontal='center')
        cell_label.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))

        cell_val = ws.cell(row=row, column=3, value=val)
        cell_val.font = Font(bold=True, size=14, color='2E86AB')
        cell_val.alignment = Alignment(horizontal='center')
        cell_val.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Monthly comparison table - row 12+
    row = 12
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value='Monthly Summary').font = Font(bold=True, size=12)
    row += 1

    month_headers = ['Metric'] + months + ['Total']
    for col_idx, header in enumerate(month_headers, 1):
        write_header(ws, row, col_idx, header, fill=fill)

    # Metric rows
    metrics = [
        ('Total Calls', []),
        ('Successful', []),
        ('Failed', []),
        ('Abandoned', []),
        ('Success Rate', []),
        ('Abandon Rate', []),
        ('Avg Duration', []),
    ]

    last_col = 1
    for metric_name, vals in metrics:
        if metric_name == 'Total Calls':
            for m in months:
                for r in outcomes:
                    if r.get('Month', '').strip() == m and r.get('Total'):
                        vals.append(int(r['Total']))
                        break
            total_val = sum(vals)
        elif metric_name == 'Successful':
            for m in months:
                for r in outcomes:
                    if r.get('Month', '').strip() == m and r.get('Successful'):
                        vals.append(int(r['Successful']))
                        break
            total_val = sum(vals)
        elif metric_name == 'Failed':
            for m in months:
                for r in outcomes:
                    if r.get('Month', '').strip() == m and r.get('Failed'):
                        vals.append(int(r['Failed']))
                        break
            total_val = sum(vals)
        elif metric_name == 'Abandoned':
            for m in months:
                for r in outcomes:
                    if r.get('Month', '').strip() == m and r.get('Abandoned'):
                        vals.append(int(r['Abandoned']))
                        break
            total_val = sum(vals)
        elif metric_name == 'Success Rate':
            for m in months:
                for r in outcomes:
                    if r.get('Month', '').strip() == m and r.get('Success%'):
                        vals.append(r['Success%'])
                        break
            total_val = ""
        elif metric_name == 'Abandon Rate':
            for m in months:
                for r in outcomes:
                    if r.get('Month', '').strip() == m and r.get('Abandon%'):
                        vals.append(r['Abandon%'])
                        break
            total_val = ""
        elif metric_name == 'Avg Duration':
            for m in months:
                for r in duration:
                    if r.get('Month', '').strip() == m and r.get('Avg Duration (s)'):
                        vals.append(format_duration(float(r['Avg Duration (s)'])))
                        break
            total_val = format_duration(overall_avg)

        row += 1
        write_cell(ws, row, 1, metric_name, bold=True)
        for col_idx, val in enumerate(vals, 2):
            write_cell(ws, row, col_idx, val)
        write_cell(ws, row, col_idx + 1, total_val)

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['E'].width = 12


def create_monthly_tab(wb, month_name, outcomes, duration, flow_paths_month, summary):
    """Create a monthly tab with Key Metrics and Top 10 flow paths."""
    ws = wb.create_sheet(title=month_name)

    # Key Metrics
    row = 1
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value=f'{month_name} 2026 - Key Metrics').font = Font(bold=True, size=12)
    row += 1

    headers = ['Metric', 'Value']
    for col_idx, h in enumerate(headers, 1):
        write_header(ws, row, col_idx, h, fill=PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid'))

    # Find this month's data
    month_outcome = None
    month_dur = None
    for r in outcomes:
        if r.get('Month', '').strip() == month_name:
            month_outcome = r
            break
    for r in duration:
        if r.get('Month', '').strip() == month_name:
            month_dur = r
            break

    metrics = [
        ('Total Calls', month_outcome.get('Total', '0') if month_outcome else '0'),
        ('Successful', month_outcome.get('Successful', '0') if month_outcome else '0'),
        ('Failed', month_outcome.get('Failed', '0') if month_outcome else '0'),
        ('Abandoned', month_outcome.get('Abandoned', '0') if month_outcome else '0'),
        ('Success Rate', f'{month_outcome.get("Success%", "0")}%' if month_outcome else '0%'),
        ('Abandon Rate', f'{month_outcome.get("Abandon%", "0")}%' if month_outcome else '0%'),
        ('Avg Duration', month_dur.get('Avg Duration (s)', '0') if month_dur else '0'),
        ('P50 Duration', month_dur.get('P50 Duration (s)', '0') if month_dur else '0'),
        ('P95 Duration', month_dur.get('P95 Duration (s)', '0') if month_dur else '0'),
    ]

    for metric_name, val in metrics:
        row += 1
        if 'Duration' in metric_name and metric_name not in ('Avg Duration', 'P50 Duration', 'P95 Duration', 'Max Duration'):
            pass
        elif 'Duration' in metric_name:
            try:
                dur_val = float(val)
                val = format_duration(dur_val)
            except (ValueError, TypeError):
                pass
        elif isinstance(val, str) and val.endswith('%'):
            pass
        elif isinstance(val, str) and val.replace('.', '').isdigit():
            try:
                val = int(float(val))
            except (ValueError, TypeError):
                pass
        write_cell(ws, row, 1, metric_name, bold=True)
        write_cell(ws, row, 2, val)

    # Top 10 Call Flow Paths - row 18
    row = 18
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=1, value='Top 10 Call Flow Paths').font = Font(bold=True, size=12)
    row += 1

    path_headers = ['Rank', 'Path', 'Count']
    for col_idx, h in enumerate(path_headers, 1):
        write_header(ws, row, col_idx, h, fill=PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid'))

    # Read flow paths
    paths = []
    if flow_paths_month:
        for p in flow_paths_month:
            try:
                paths.append((p.get('Path', ''), int(p.get('Count', 0))))
            except (ValueError, TypeError):
                pass
    paths.sort(key=lambda x: x[1], reverse=True)

    for i, (path, count) in enumerate(paths[:10]):
        row += 1
        write_cell(ws, row, 1, i + 1)
        write_cell(ws, row, 2, path, wrap=True)
        write_cell(ws, row, 3, count)

    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 10


def create_transfers_tab(wb, transfers):
    """Create Transfers tab with transfer counts and rates per month."""
    ws = wb.create_sheet(title="Transfers")

    row = 1
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value='Transfer Patterns').font = Font(bold=True, size=12)
    row += 1

    headers = ['Month', 'Total Transfers', 'Blind Transfers', 'Consult Transfers', 'Transfer Rate (%)']
    for col_idx, h in enumerate(headers, 1):
        write_header(ws, row, col_idx, h, fill=PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid'))

    for r in transfers:
        row += 1
        write_cell(ws, row, 1, r.get('Month', ''))
        write_cell(ws, row, 2, r.get('Total Transfers', 0))
        write_cell(ws, row, 3, r.get('Blind Transfers', 0))
        write_cell(ws, row, 4, r.get('Consult Transfers', 0))
        write_cell(ws, row, 5, r.get('Transfer Rate (%)', 0))


def create_disconnect_tab(wb, disconnects):
    """Create Disconnect Types tab with top 15 disconnect types."""
    ws = wb.create_sheet(title="Disconnect Types")

    row = 1
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value='Top 15 Disconnect Types').font = Font(bold=True, size=12)
    row += 1

    headers = ['Rank', 'Disconnect Type', 'Details']
    for col_idx, h in enumerate(headers, 1):
        write_header(ws, row, col_idx, h, fill=PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid'))

    for i, r in enumerate(disconnects[:15]):
        row += 1
        write_cell(ws, row, 1, i + 1)
        write_cell(ws, row, 2, r.get('Disconnect Type', ''), wrap=True)
        # Combine all month data into details
        details_parts = []
        for key, val in r.items():
            if key not in ('Rank', 'Disconnect Type', 'Total') and str(val).strip():
                details_parts.append(f"{key}: {val}")
        write_cell(ws, row, 3, '; '.join(details_parts), wrap=True)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 60


def create_hourly_tab(wb, hourly):
    """Create Hourly tab with call distribution by hour."""
    ws = wb.create_sheet(title="Hourly")

    row = 1
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value='Hourly Call Distribution').font = Font(bold=True, size=12)
    row += 1

    # Get all month headers from first row
    months = []
    if hourly:
        for key in hourly[0].keys():
            if key not in ('Hour', 'All'):
                months.append(key)

    headers = ['Hour'] + months + ['All']
    for col_idx, h in enumerate(headers, 1):
        write_header(ws, row, col_idx, h, fill=PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid'))

    for r in hourly:
        row += 1
        hour = int(r.get('Hour', 0))
        write_cell(ws, row, 1, hour)
        for col_idx, month in enumerate(months, 2):
            write_cell(ws, row, col_idx, r.get(month, 0))
        write_cell(ws, row, col_idx + 1, r.get('All', 0))

    ws.column_dimensions['A'].width = 8
    for i, _ in enumerate(months):
        col_letter = chr(66 + i)  # B, C, D...
        ws.column_dimensions[col_letter].width = 15


def create_duration_tab(wb, duration):
    """Create Duration Statistics tab."""
    ws = wb.create_sheet(title="Duration")

    row = 1
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value='Duration Statistics').font = Font(bold=True, size=12)
    row += 1

    headers = ['Month', 'Avg Duration', 'P50 Duration', 'P95 Duration', 'Max Duration', 'Valid Samples']
    for col_idx, h in enumerate(headers, 1):
        write_header(ws, row, col_idx, h, fill=PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid'))

    for r in duration:
        row += 1
        write_cell(ws, row, 1, r.get('Month', ''))
        for col_idx, metric in enumerate(['Avg Duration (s)', 'P50 Duration (s)', 'P95 Duration (s)', 'Max Duration (s)'], 2):
            val = r.get(metric, '0')
            try:
                val = format_duration(float(val))
            except (ValueError, TypeError):
                pass
            write_cell(ws, row, col_idx, val)
        write_cell(ws, row, col_idx + 1, r.get('Valid Samples', 0))


def create_outcomes_tab(wb, outcomes):
    """Create Outcomes Stacked tab."""
    ws = wb.create_sheet(title="Outcomes Stacked")

    row = 1
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value='Call Outcomes by Month').font = Font(bold=True, size=12)
    row += 1

    headers = ['Month', 'Total', 'Successful', 'Failed', 'Abandoned']
    for col_idx, h in enumerate(headers, 1):
        write_header(ws, row, col_idx, h, fill=PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid'))

    for r in outcomes:
        row += 1
        write_cell(ws, row, 1, r.get('Month', ''))
        write_cell(ws, row, 2, r.get('Total', 0))
        write_cell(ws, row, 3, r.get('Successful', 0))
        write_cell(ws, row, 4, r.get('Failed', 0))
        write_cell(ws, row, 5, r.get('Abandoned', 0))


def create_flow_paths_tab(wb, flow_paths):
    """Create Flow Paths tab with all unique paths."""
    ws = wb.create_sheet(title="Flow Paths")

    row = 1
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=1, value='All Call Flow Paths').font = Font(bold=True, size=12)
    row += 1

    headers = ['Path', 'Count']
    for col_idx, h in enumerate(headers, 1):
        write_header(ws, row, col_idx, h, fill=PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid'))

    for path, count in flow_paths:
        row += 1
        write_cell(ws, row, 1, path, wrap=True)
        write_cell(ws, row, 2, count)

    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 12


def generate_report(category=None):
    """Generate the complete Excel report."""
    if category:
        category_name = category
    else:
        # Auto-detect from available CSV files
        available = []
        for fname in os.listdir(CHARTS_DIR):
            if fname.endswith('_outcomes.csv'):
                available.append(fname.replace('_outcomes.csv', ''))
        if available:
            category_name = available[0]
        else:
            print("No data CSVs found in Reports/Charts/. Run extract_html_data.py first.")
            return

    print(f"Generating report for category: {category_name}")

    # Read CSV data
    outcomes = read_csv_file(os.path.join(CHARTS_DIR, f'{category_name}_outcomes.csv'))
    duration = read_csv_file(os.path.join(CHARTS_DIR, f'{category_name}_duration.csv'))
    hourly = read_csv_file(os.path.join(CHARTS_DIR, f'{category_name}_hourly.csv'))
    transfers = read_csv_file(os.path.join(CHARTS_DIR, f'{category_name}_transfers.csv'))
    disconnects = read_csv_file(os.path.join(CHARTS_DIR, f'{category_name}_disconnect_types.csv'))
    flow_paths_all = read_csv_file(os.path.join(CHARTS_DIR, f'{category_name}_flow_paths.csv'))

    # Build flow paths list (sorted by count)
    flow_paths_list = []
    for r in flow_paths_all:
        try:
            flow_paths_list.append((r.get('Path', ''), int(r.get('Count', 0))))
        except (ValueError, TypeError):
            pass
    flow_paths_list.sort(key=lambda x: x[1], reverse=True)

    # Read monthly flow paths
    months = get_ordered_months(outcomes)
    monthly_flow_paths = {}
    for month in months:
        monthly_csv = os.path.join(CHARTS_DIR, f'{category_name}_flow_paths_{month.lower()}.csv')
        mp = read_csv_file(monthly_csv)
        monthly_flow_paths[month] = mp

    # Create workbook
    wb = Workbook()

    # Create Summary tab first (active)
    ws = wb.active
    ws.title = "Summary"

    # Create Summary tab first (active)
    create_summary_tab(wb, outcomes, duration, transfers, flow_paths_list)

    # Create monthly tabs
    for month in months:
        create_monthly_tab(wb, month, outcomes, duration, monthly_flow_paths.get(month, []), None)

    # Create remaining tabs
    create_transfers_tab(wb, transfers)
    create_disconnect_tab(wb, disconnects)
    create_hourly_tab(wb, hourly)
    create_duration_tab(wb, duration)
    create_outcomes_tab(wb, outcomes)
    create_flow_paths_tab(wb, flow_paths_list)

    # Write output
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    months_short = '-'.join(m[:3] for m in months)
    output_name = f'{category_name}_DeepDive_{months_short}_2026.xlsx'
    output_path = os.path.join(OUTPUT_DIR, output_name)
    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"  Tabs: Summary, {', '.join(months)}, Transfers, Disconnect Types, Hourly, Duration, Outcomes Stacked, Flow Paths")


if __name__ == '__main__':
    cat = sys.argv[1] if len(sys.argv) > 1 else None
    generate_report(cat)