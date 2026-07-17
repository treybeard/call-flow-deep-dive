#!/usr/bin/env python3
"""Convert Switchboard Deep Dive XLSX to HTML with charts and graphs."""

import os
import sys
import io
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

sys.path.insert(0, '/Library/Python/3.9/lib/python/site-packages')
from openpyxl import load_workbook

PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(PROJECT_DIR, "Output", "Switchboard_DeepDive_Apr-Jun_2026.xlsx")
OUTPUT_DIR = os.path.join(PROJECT_DIR, "Reports_HTML")
CHARTS_DIR = os.path.join(OUTPUT_DIR, "charts")

COLORS = {
    'primary': '#2E86AB',
    'secondary': '#A23B72',
    'success': '#28a745',
    'danger': '#dc3545',
    'warning': '#ffc107',
    'info': '#17a2b8',
}

def parse_duration(dur_str):
    if ':' in str(dur_str):
        parts = str(dur_str).split(':')
        return int(parts[0]) * 60 + int(parts[1])
    return float(dur_str)

def create_summary_charts(ws):
    charts = []
    # KPI cards
    kpis = {}
    for row in ws.iter_rows(min_row=4, max_row=8, min_col=1, max_col=3, values_only=True):
        if row[0] and row[1]:
            kpis[str(row[0])] = str(row[1])
    
    fig, axes = plt.subplots(2, 2, figsize=(12, 8))
    kpi_labels = list(kpis.keys())[:4]
    kpi_values = list(kpis.values())[:4]
    for i, ax in enumerate(axes.flat):
        if i < len(kpi_labels):
            ax.axis('off')
            ax.text(0.5, 0.5, f'{kpi_values[i]}', ha='center', va='center', fontsize=32, fontweight='bold', color=COLORS['primary'])
            ax.text(0.5, 0.2, kpi_labels[i], ha='center', va='center', fontsize=14, color='#666')
    plt.suptitle('Key Performance Indicators', fontsize=16, fontweight='bold')
    plt.tight_layout()
    chart_path = os.path.join(CHARTS_DIR, 'summary_kpis.png')
    fig.savefig(chart_path, bbox_inches='tight', dpi=150, facecolor='white')
    plt.close(fig)
    charts.append(('summary_kpis.png', 'KPI Cards'))
    
    # Monthly summary comparison
    monthly_data = {}
    for row in ws.iter_rows(min_row=21, max_row=27, min_col=1, max_col=5, values_only=True):
        if row[0]:
            monthly_data[row[0]] = [row[i] if row[i] else 0 for i in range(1, 5)]
    
    for key in monthly_data:
        for i, val in enumerate(monthly_data[key]):
            if isinstance(val, str) and '%' in val:
                monthly_data[key][i] = float(val.replace('%', '')) / 100
    
    months = ['April', 'May', 'June']
    metrics = ['Total Calls', 'Successful', 'Failed', 'Abandoned', 'Success Rate', 'Abandon Rate', 'Avg Duration']
    
    fig, ax = plt.subplots(figsize=(12, 7))
    x = range(len(months))
    width = 0.15
    
    for i, metric in enumerate(metrics):
        if metric in monthly_data:
            values = monthly_data[metric][:3]
            offset = (i - 2.5) * width
            bar = ax.bar([p + offset for p in x], values, width, label=metric)
    
    ax.set_xlabel('Month')
    ax.set_ylabel('Value')
    ax.set_title('Monthly Summary Comparison', fontsize=14, fontweight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(months)
    ax.legend(loc='upper right', fontsize=9)
    
    chart_path = os.path.join(CHARTS_DIR, 'monthly_summary.png')
    fig.savefig(chart_path, bbox_inches='tight', dpi=150, facecolor='white')
    plt.close(fig)
    charts.append(('monthly_summary.png', 'Monthly Summary Comparison'))
    
    return charts

def create_monthly_charts(ws, month_name):
    charts = []
    ws_data = {}
    for row in ws.iter_rows(min_row=2, max_row=40, min_col=1, max_col=3, values_only=True):
        if row[0] and row[1]:
            ws_data[str(row[0])] = row[1:]
    
    metrics = {}
    if 'Key Metrics' in ws_data:
        key_vals = ws_data['Key Metrics']
        for val in key_vals:
            if isinstance(val, tuple) and len(val) >= 2:
                k = str(val[0]).strip() if val[0] else ''
                v = str(val[1]).strip() if val[1] else ''
                if k:
                    metrics[k] = v
    
    # Call distribution pie chart
    data_section = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5, values_only=True):
        if row[0] and 'Calls' in str(row[0]).lower():
            data_section = row
            break
    
    if data_section:
        labels = []
        values = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4, values_only=True):
            if row[0] and row[1] and isinstance(row[0], str):
                if row[1]:
                    labels.append(str(row[0]))
                    try:
                        values.append(int(float(str(row[1]) if row[1] else '0')))
                    except (ValueError, TypeError):
                        pass
    
    if labels:
        fig, ax = plt.subplots(figsize=(10, 8))
        ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90, colors=[COLORS['primary'], COLORS['secondary'], COLORS['info'], COLORS['warning'], COLORS['success']])
        ax.set_title(f'{month_name} 2026 - Call Distribution', fontsize=14, fontweight='bold')
        chart_path = os.path.join(CHARTS_DIR, f'{month_name.lower()}_call_distribution.png')
        fig.savefig(chart_path, bbox_inches='tight', dpi=150, facecolor='white')
        plt.close(fig)
        charts.append((f'{month_name.lower()}_call_distribution.png', f'{month_name} 2026 - Call Distribution'))
    
    # Duration comparison bar chart
    avg_duration = metrics.get('Avg Duration', '06:26')
    p50_duration = metrics.get('P50 Duration', '05:10')
    p95_duration = metrics.get('P95 Duration', '16:41')
    
    durations = [parse_duration(avg_duration), parse_duration(p50_duration), parse_duration(p95_duration)]
    duration_labels = ['Avg Duration', 'P50 Duration', 'P95 Duration']
    
    fig, ax = plt.subplots(figsize=(8, 6))
    bars = ax.bar(duration_labels, durations, color=[COLORS['info'], COLORS['primary'], COLORS['secondary']], edgecolor='white', linewidth=2)
    ax.set_ylabel('Duration (seconds)')
    ax.set_title(f'{month_name} 2026 - Duration Statistics', fontsize=14, fontweight='bold')
    ax.grid(axis='y', alpha=0.3)
    
    for bar, duration in zip(bars, durations):
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width() / 2., height,
                f'{int(duration // 60):02d}:{int(duration % 60):02d}',
                ha='center', va='bottom', fontweight='bold', fontsize=11)
    
    chart_path = os.path.join(CHARTS_DIR, f'{month_name.lower()}_duration.png')
    fig.savefig(chart_path, bbox_inches='tight', dpi=150, facecolor='white')
    plt.close(fig)
    charts.append((f'{month_name.lower()}_duration.png', f'{month_name} Duration Statistics'))
    
    # Top call paths
    paths_data = []
    for row in ws.iter_rows(min_row=19, max_row=28, min_col=1, max_col=3, values_only=True):
        if row[0] and row[1]:
            paths_data.append({
                'rank': int(row[0]),
                'path': str(row[1]),
                'count': int(row[2]) if row[2] else 0
            })
    
    if paths_data:
        top_paths = paths_data[:10]
        
        fig, ax = plt.subplots(figsize=(10, 6))
        y_pos = range(len(top_paths))
        counts = [p['count'] for p in top_paths]
        
        bars = ax.barh(y_pos, counts, color=COLORS['primary'], edgecolor='white', linewidth=1)
        ax.set_yticks(y_pos)
        ax.set_yticklabels([f"#{p['rank']} {p['path'][:40]}" for p in top_paths], fontsize=9)
        ax.set_xlabel('Number of Calls')
        ax.set_title(f'Top 10 Call Flow Paths - {month_name} 2026', fontsize=14, fontweight='bold')
        ax.grid(axis='x', alpha=0.3)
        
        for i, (bar, count) in enumerate(zip(bars, counts)):
            ax.text(bar.get_width() + max(counts) * 0.01, bar.get_y() + bar.get_height() / 2,
                    f'{count:,}', ha='left', va='center', fontweight='bold', fontsize=10)
        
        chart_path = os.path.join(CHARTS_DIR, f'{month_name.lower()}_top_paths.png')
        fig.savefig(chart_path, bbox_inches='tight', dpi=150, facecolor='white')
        plt.close(fig)
        charts.append((f'{month_name.lower()}_top_paths.png', f'{month_name} Top Call Flow Paths'))
    
    return charts

def create_transfers_charts(ws):
    charts = []
    transfer_data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4, values_only=True):
        if row[0] and row[1]:
            transfer_data.append({
                'month': str(row[0]),
                'count': int(row[1]) if row[1] else 0,
                'rate': float(str(row[2]).replace('%', '')) / 100 if row[2] and '%' in str(row[2]) else (float(row[2]) if row[2] else 0)
            })
    
    if transfer_data:
        fig, ax1 = plt.subplots(figsize=(10, 6))
        
        months = [d['month'] for d in transfer_data]
        counts = [d['count'] for d in transfer_data]
        rates = [d['rate'] for d in transfer_data]
        
        x = range(len(months))
        width = 0.35
        
        bars = ax1.bar([p + width/2 for p in x], counts, width, label='Transfer Count', color=COLORS['primary'])
        ax2 = ax1.twinx()
        ax2.plot([p + width/2 for p in x], rates, 'o-', color=COLORS['danger'], linewidth=2, label='Transfer Rate')
        
        ax1.set_xlabel('Month')
        ax1.set_ylabel('Transfer Count', color=COLORS['primary'])
        ax2.set_ylabel('Transfer Rate', color=COLORS['danger'])
        ax1.set_title('Transfer Patterns', fontsize=14, fontweight='bold')
        ax1.set_xticks([p + width/2 for p in x])
        ax1.set_xticklabels(months)
        
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper right')
        ax1.grid(alpha=0.3)
        
        chart_path = os.path.join(CHARTS_DIR, 'transfers.png')
        fig.savefig(chart_path, bbox_inches='tight', dpi=150, facecolor='white')
        plt.close(fig)
        charts.append(('transfers.png', 'Transfer Patterns'))
    
    return charts

def create_disconnect_charts(ws):
    charts = []
    disconnect_data = []
    for row in ws.iter_rows(min_row=4, max_row=18, min_col=1, max_col=3, values_only=True):
        if row[0] and row[1]:
            disconnect_data.append({
                'rank': int(row[0]),
                'type': str(row[1]),
                'total': int(row[2]) if row[2] else 0
            })
    
    if disconnect_data:
        top_disconnected = disconnect_data[:10]
        
        fig, ax = plt.subplots(figsize=(14, 8))
        y_pos = range(len(top_disconnected))
        counts = [d['total'] for d in top_disconnected]
        
        bars = ax.barh(y_pos, counts, color=COLORS['danger'], edgecolor='white', linewidth=1, height=0.6)
        ax.set_yticks(y_pos)
        ax.set_yticklabels([d['type'][:55] + '...' if len(d['type']) > 55 else d['type'] for d in top_disconnected], fontsize=10)
        ax.set_xlabel('Number of Calls')
        ax.set_title('Top 10 Disconnect Types', fontsize=14, fontweight='bold')
        ax.grid(axis='x', alpha=0.3)
        
        for bar, count in zip(bars, counts):
            ax.text(bar.get_width() + max(counts) * 0.01, bar.get_y() + bar.get_height() / 2,
                    f'{count:,}', ha='left', va='center', fontweight='bold', fontsize=11)
        
        chart_path = os.path.join(CHARTS_DIR, 'disconnect_types.png')
        fig.savefig(chart_path, bbox_inches='tight', dpi=150, facecolor='white')
        plt.close(fig)
        charts.append(('disconnect_types.png', 'Disconnect Types'))
    
    return charts

def create_hourly_charts(ws):
    charts = []
    hours = []
    april_data = []
    may_data = []
    june_data = []
    
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=4, values_only=True):
        if row[0] is not None and isinstance(row[0], (int, float)):
            hours.append(int(row[0]))
            april_data.append(int(row[1]) if row[1] else 0)
            may_data.append(int(row[2]) if row[2] else 0)
            june_data.append(int(row[3]) if row[3] else 0)
    
    if hours:
        fig, ax = plt.subplots(figsize=(12, 6))
        ax.plot(hours, april_data, marker='o', linewidth=2, label='April', color=COLORS['primary'])
        ax.plot(hours, may_data, marker='s', linewidth=2, label='May', color=COLORS['secondary'])
        ax.plot(hours, june_data, marker='^', linewidth=2, label='June', color=COLORS['info'])
        ax.set_xlabel('Hour of Day')
        ax.set_ylabel('Number of Calls')
        ax.set_title('Hourly Call Distribution', fontsize=14, fontweight='bold')
        ax.set_xticks(hours)
        ax.legend()
        ax.grid(alpha=0.3)
        
        chart_path = os.path.join(CHARTS_DIR, 'hourly_distribution.png')
        fig.savefig(chart_path, bbox_inches='tight', dpi=150, facecolor='white')
        plt.close(fig)
        charts.append(('hourly_distribution.png', 'Hourly Distribution'))
    
    return charts

def create_duration_stats_charts(ws):
    charts = []
    duration_data = []
    for row in ws.iter_rows(min_row=4, max_row=8, min_col=1, max_col=5, values_only=True):
        if row[0]:
            entry = {'metric': str(row[0])}
            for i in range(1, 5):
                if row[i]:
                    try:
                        entry[f'month{i-1}'] = parse_duration(str(row[i]))
                    except:
                        entry[f'month{i-1}'] = 0
            duration_data.append(entry)
    
    if duration_data:
        metrics = [d['metric'] for d in duration_data]
        avgs = [d.get('month0', 0) for d in duration_data]
        p50s = [d.get('month1', 0) for d in duration_data]
        p95s = [d.get('month2', 0) for d in duration_data]
        maxes = [d.get('month3', 0) for d in duration_data]
        
        x = range(len(metrics))
        width = 0.2
        
        fig, ax = plt.subplots(figsize=(12, 7))
        ax.bar([p + 0*width for p in x], avgs, width, label='Avg', color=COLORS['primary'])
        ax.bar([p + 1*width for p in x], p50s, width, label='P50', color=COLORS['info'])
        ax.bar([p + 2*width for p in x], p95s, width, label='P95', color=COLORS['secondary'])
        ax.bar([p + 3*width for p in x], maxes, width, label='Max', color=COLORS['warning'])
        
        ax.set_xlabel('Metric')
        ax.set_ylabel('Duration (seconds)')
        ax.set_title('Duration Statistics by Month', fontsize=14, fontweight='bold')
        ax.set_xticks([p + 1.5*width for p in x])
        ax.set_xticklabels(metrics)
        ax.legend()
        ax.grid(axis='y', alpha=0.3)
        
        chart_path = os.path.join(CHARTS_DIR, 'duration_stats.png')
        fig.savefig(chart_path, bbox_inches='tight', dpi=150, facecolor='white')
        plt.close(fig)
        charts.append(('duration_stats.png', 'Duration Statistics'))
    
    return charts

def create_outcomes_charts(ws):
    charts = []
    months = []
    totals = []
    successful = []
    failed = []
    abandoned = []
    
    for row in ws.iter_rows(min_row=4, max_row=6, min_col=1, max_col=5, values_only=True):
        if row[0]:
            months.append(str(row[0]))
            totals.append(int(row[1]) if row[1] else 0)
            successful.append(int(row[2]) if row[2] else 0)
            failed.append(int(row[3]) if row[3] else 0)
            abandoned.append(int(row[4]) if row[4] else 0)
    
    if months:
        fig, ax = plt.subplots(figsize=(10, 6))
        x = range(len(months))
        width = 0.6
        
        bars1 = ax.bar(x, successful, width, label='Successful', color=COLORS['success'])
        bars2 = ax.bar(x, failed, width, bottom=successful, label='Failed', color=COLORS['danger'])
        bars3 = ax.bar(x, abandoned, width, bottom=[s+f for s, f in zip(successful, failed)], label='Abandoned', color=COLORS['warning'])
        
        ax.set_xlabel('Month')
        ax.set_ylabel('Number of Calls')
        ax.set_title('Call Outcomes', fontsize=14, fontweight='bold')
        ax.set_xticks(x)
        ax.set_xticklabels(months)
        ax.legend()
        ax.grid(axis='y', alpha=0.3)
        
        chart_path = os.path.join(CHARTS_DIR, 'outcomes.png')
        fig.savefig(chart_path, bbox_inches='tight', dpi=150, facecolor='white')
        plt.close(fig)
        charts.append(('outcomes.png', 'Call Outcomes'))
    
    return charts

def create_flow_paths_charts(ws):
    charts = []
    flow_data = []
    for row in ws.iter_rows(min_row=4, max_row=20, min_col=1, max_col=3, values_only=True):
        if row[0] and row[1]:
            flow_data.append({
                'path': str(row[0]),
                'count': int(row[1]) if row[1] else 0
            })
    
    if flow_data:
        flow_data.sort(key=lambda x: x['count'], reverse=True)
        top_flow_paths = flow_data[:15]
        
        fig, ax = plt.subplots(figsize=(12, 8))
        y_pos = range(len(top_flow_paths))
        counts = [p['count'] for p in top_flow_paths]
        
        bars = ax.barh(y_pos, counts, color=COLORS['secondary'], edgecolor='white', linewidth=1)
        ax.set_yticks(y_pos)
        ax.set_yticklabels([p['path'][:60] + '...' if len(p['path']) > 60 else p['path'] for p in top_flow_paths], fontsize=9)
        ax.set_xlabel('Number of Calls')
        ax.set_title('Top 15 Call Flow Paths', fontsize=14, fontweight='bold')
        ax.grid(axis='x', alpha=0.3)
        
        for bar, count in zip(bars, counts):
            ax.text(bar.get_width() + max(counts) * 0.01, bar.get_y() + bar.get_height() / 2,
                    f'{count:,}', ha='left', va='center', fontweight='bold', fontsize=10)
        
        chart_path = os.path.join(CHARTS_DIR, 'flow_paths.png')
        fig.savefig(chart_path, bbox_inches='tight', dpi=150, facecolor='white')
        plt.close(fig)
        charts.append(('flow_paths.png', 'Call Flow Paths'))
    
    return charts

def generate_html():
    wb = load_workbook(XLSX_PATH)
    os.makedirs(CHARTS_DIR, exist_ok=True)
    
    all_charts = []
    all_charts.extend(create_summary_charts(wb['Summary']))
    
    for month in ['April', 'May', 'June']:
        all_charts.extend(create_monthly_charts(wb[month], month))
    
    all_charts.extend(create_transfers_charts(wb['Transfers']))
    all_charts.extend(create_disconnect_charts(wb['Disconnect Types']))
    all_charts.extend(create_hourly_charts(wb['Hourly']))
    all_charts.extend(create_duration_stats_charts(wb['Duration']))
    all_charts.extend(create_outcomes_charts(wb['Outcomes Stacked']))
    all_charts.extend(create_flow_paths_charts(wb['Flow Paths']))
    
    charts_html = ""
    for chart_file, chart_name in all_charts:
        charts_html += f"""
        <div class="chart-card">
            <h3>{chart_name}</h3>
            <img src="charts/{chart_file}" alt="{chart_name}" class="chart-image">
        </div>
        """
    
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Switchboard Deep-Dive Analysis - April to June 2026</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: #F8F9FA;
            color: #2C3E50;
            line-height: 1.6;
        }}
        .container {{ max-width: 1400px; margin: 0 auto; padding: 20px; }}
        header {{
            background: linear-gradient(135deg, #2E86AB, #A23B72);
            color: white;
            padding: 40px 0;
            text-align: center;
            margin-bottom: 40px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }}
        header h1 {{ font-size: 2.5em; margin-bottom: 10px; font-weight: 300; }}
        header p {{ font-size: 1.2em; opacity: 0.9; }}
        .dashboard {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-bottom: 40px;
        }}
        .kpi-card {{
            background: white;
            border-radius: 8px;
            padding: 20px;
            text-align: center;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .kpi-value {{ font-size: 2em; font-weight: bold; color: #2E86AB; }}
        .kpi-label {{ color: #666; margin-top: 5px; }}
        .chart-card {{
            background: white;
            border-radius: 8px;
            padding: 20px;
            margin-bottom: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .chart-card h3 {{
            margin-bottom: 15px;
            color: #2C3E50;
            border-bottom: 2px solid #2E86AB;
            padding-bottom: 10px;
        }}
        .chart-image {{
            width: 100%;
            max-width: 1000px;
            display: block;
            margin: 0 auto;
        }}
        .section {{ margin-bottom: 40px; }}
        .section h2 {{
            color: #2E86AB;
            margin-bottom: 20px;
            font-size: 1.8em;
        }}
        .charts-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(600px, 1fr));
            gap: 20px;
        }}
        footer {{
            text-align: center;
            padding: 20px;
            color: #666;
            border-top: 1px solid #ddd;
            margin-top: 40px;
        }}
    </style>
</head>
<body>
    <header>
        <div class="container">
            <h1>Switchboard Deep-Dive Analysis</h1>
            <p>April to June 2026</p>
        </div>
    </header>
    <div class="container">
        <div class="dashboard">
            <div class="kpi-card">
                <div class="kpi-value">8,620</div>
                <div class="kpi-label">Total Calls</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-value">14.8%</div>
                <div class="kpi-label">Abandonment Rate</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-value">66.9%</div>
                <div class="kpi-label">Success Rate</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-value">06:40</div>
                <div class="kpi-label">Avg Duration</div>
            </div>
        </div>
        <div class="section">
            <h2>Summary</h2>
            <div class="charts-grid">
                {charts_html}
            </div>
        </div>
        <div class="section">
            <h2>Detailed Analysis</h2>
            <div class="charts-grid">
                {charts_html}
            </div>
        </div>
    </div>
    <footer>
        <div class="container">
            <p>Generated by Call Flow Deep-Dive Analysis Tool</p>
        </div>
    </footer>
</body>
</html>"""
    
    html_path = os.path.join(OUTPUT_DIR, 'report_Switchboard_2026.html')
    with open(html_path, 'w') as f:
        f.write(html)
    
    print(f"Generated: {html_path}")
    return html_path

if __name__ == '__main__':
    generate_html()
